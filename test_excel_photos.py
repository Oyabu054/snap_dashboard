# -*- coding: utf-8 -*-
import json
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

import excel_photos


def _write_index(path, entries):
    path.write_text(json.dumps(entries, ensure_ascii=False), encoding="utf-8")


def _png_bytes(color=(255, 0, 0)):
    buf = BytesIO()
    PILImage.new("RGB", (2, 2), color=color).save(buf, format="PNG")
    buf.seek(0)
    return buf.read()


def _build_workbook(sheets_images):
    """
    sheets_images: {sheet_name: [image_bytes, ...]}
    openpyxlで画像を埋め込んだワークブックを組み立て、load_workbookで
    読み直した状態(実際のBoxダウンロード後と同じ状態)で返す。
    """
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, images in sheets_images.items():
        ws = wb.create_sheet(sheet_name)
        for i, img_bytes in enumerate(images):
            img = Image(BytesIO(img_bytes))
            img.anchor = f"A{i + 1}"
            ws.add_image(img)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return load_workbook(out)


def test_parse_filename_extracts_line_name_and_date():
    result = excel_photos.parse_filename("(L1)2026.03.10_0852_A品種_T5.xlsx")
    assert result == {
        "line_name": "L1",
        "date": date(2026, 3, 10),
        "filename": "(L1)2026.03.10_0852_A品種_T5.xlsx",
    }


def test_parse_filename_accepts_fullwidth_parentheses():
    result = excel_photos.parse_filename("（L1）2026.03.10_0852_A品種_T5.xlsx")
    assert result["line_name"] == "L1"
    assert result["date"] == date(2026, 3, 10)


def test_parse_filename_ignores_suffix_content():
    result = excel_photos.parse_filename("(製板2号機)2026.12.25_1730_特殊品種_厚み5mm_補足情報.xlsx")
    assert result["line_name"] == "製板2号機"
    assert result["date"] == date(2026, 12, 25)


def test_parse_filename_accepts_legacy_xls_extension():
    result = excel_photos.parse_filename("(L1)2026.03.10_0852_A品種_T5.xls")
    assert result["date"] == date(2026, 3, 10)


def test_parse_filename_returns_none_for_unmatched_name():
    assert excel_photos.parse_filename("invalid_name.xlsx") is None
    assert excel_photos.parse_filename("IMG_20260310.jpg") is None


def test_parse_filename_returns_none_for_invalid_date():
    assert excel_photos.parse_filename("(L1)2026.13.40_0852_A品種_T5.xlsx") is None


def test_get_photos_returns_entries_within_range(tmp_path):
    index_path = tmp_path / "index.json"
    _write_index(index_path, [
        {"id": "a1", "name": "a.png", "date": "2026-03-09", "line_name": "L1"},
        {"id": "a2", "name": "b.png", "date": "2026-03-10", "line_name": "L1"},
        {"id": "a3", "name": "c.png", "date": "2026-03-15", "line_name": "L1"},
    ])

    result = excel_photos.get_photos(
        datetime(2026, 3, 10), datetime(2026, 3, 12), index_path=index_path
    )

    assert [r["id"] for r in result] == ["a2"]


def test_get_photos_includes_boundary_dates(tmp_path):
    index_path = tmp_path / "index.json"
    _write_index(index_path, [
        {"id": "a1", "name": "a.png", "date": "2026-03-10", "line_name": "L1"},
        {"id": "a2", "name": "b.png", "date": "2026-03-12", "line_name": "L1"},
    ])

    result = excel_photos.get_photos(
        datetime(2026, 3, 10), datetime(2026, 3, 12), index_path=index_path
    )

    assert [r["id"] for r in result] == ["a1", "a2"]


def test_get_photos_sorted_by_timestamp(tmp_path):
    index_path = tmp_path / "index.json"
    _write_index(index_path, [
        {"id": "later", "name": "b.png", "date": "2026-03-12", "line_name": "L1"},
        {"id": "earlier", "name": "a.png", "date": "2026-03-10", "line_name": "L1"},
    ])

    result = excel_photos.get_photos(
        datetime(2026, 3, 1), datetime(2026, 3, 31), index_path=index_path
    )

    assert [r["id"] for r in result] == ["earlier", "later"]


def test_get_photos_returns_empty_list_when_index_missing(tmp_path):
    index_path = tmp_path / "does_not_exist.json"

    result = excel_photos.get_photos(
        datetime(2026, 3, 1), datetime(2026, 3, 31), index_path=index_path
    )

    assert result == []


def test_get_photos_returns_empty_list_when_index_empty(tmp_path):
    index_path = tmp_path / "index.json"
    _write_index(index_path, [])

    result = excel_photos.get_photos(
        datetime(2026, 3, 1), datetime(2026, 3, 31), index_path=index_path
    )

    assert result == []


def test_get_photos_return_shape_matches_box_client(tmp_path):
    index_path = tmp_path / "index.json"
    _write_index(index_path, [
        {"id": "a1", "name": "a.png", "date": "2026-03-10", "line_name": "L1"},
    ])

    result = excel_photos.get_photos(
        datetime(2026, 3, 1), datetime(2026, 3, 31), index_path=index_path
    )

    assert result == [{"id": "a1", "name": "a.png", "timestamp": "2026-03-10T00:00:00"}]


# ===================== _extract_images_from_workbook =====================

def test_extract_images_from_workbook_returns_sheet_bytes_format():
    png = _png_bytes()
    wb = _build_workbook({"3-10": [png]})

    result = excel_photos._extract_images_from_workbook(wb)

    assert len(result) == 1
    sheet_name, image_bytes, fmt = result[0]
    assert sheet_name == "3-10"
    assert fmt == "png"
    assert image_bytes[:8] == b"\x89PNG\r\n\x1a\n"


def test_extract_images_from_workbook_handles_multiple_sheets_and_images():
    wb = _build_workbook({
        "3-10": [_png_bytes((255, 0, 0)), _png_bytes((0, 255, 0))],
        "3-11": [_png_bytes((0, 0, 255))],
    })

    result = excel_photos._extract_images_from_workbook(wb)

    assert len(result) == 3
    assert sorted(s for s, _, _ in result) == ["3-10", "3-10", "3-11"]


def test_extract_images_from_workbook_returns_empty_list_when_no_images():
    wb = _build_workbook({"3-10": []})

    result = excel_photos._extract_images_from_workbook(wb)

    assert result == []


# ===================== sync_cache =====================

def test_sync_cache_extracts_images_and_writes_index(tmp_path):
    png = _png_bytes()
    wb = _build_workbook({"sheet1": [png]})
    out = BytesIO()
    wb.save(out)
    xlsx_bytes = out.getvalue()

    list_files = lambda: [
        {"id": "file1", "name": "(L1)2026.03.10_0852_A品種_T5.xlsx", "updated_at": "2026-03-10T08:52:00"},
    ]
    download_file = lambda file_id: xlsx_bytes

    summary = excel_photos.sync_cache(
        cache_dir=tmp_path, list_files=list_files, download_file=download_file
    )

    assert summary == {"processed": 1, "skipped": 0, "photos": 1}

    index_path = tmp_path / "index.json"
    entries = json.loads(index_path.read_text(encoding="utf-8"))
    assert len(entries) == 1
    assert entries[0]["date"] == "2026-03-10"
    assert entries[0]["line_name"] == "L1"
    assert (tmp_path / "photos" / "2026-03-10").exists()


def test_sync_cache_skips_files_with_unmatched_naming(tmp_path):
    list_files = lambda: [{"id": "file1", "name": "invalid_name.xlsx", "updated_at": "2026-03-10T08:52:00"}]
    download_file = lambda file_id: (_ for _ in ()).throw(AssertionError("マッチしないファイルはダウンロードしないはず"))

    summary = excel_photos.sync_cache(
        cache_dir=tmp_path, list_files=list_files, download_file=download_file
    )

    assert summary == {"processed": 0, "skipped": 1, "photos": 0}


def test_sync_cache_skips_unchanged_files_on_resync(tmp_path):
    png = _png_bytes()
    wb = _build_workbook({"sheet1": [png]})
    out = BytesIO()
    wb.save(out)
    xlsx_bytes = out.getvalue()

    files = [{"id": "file1", "name": "(L1)2026.03.10_0852_A品種_T5.xlsx", "updated_at": "2026-03-10T08:52:00"}]
    calls = {"download": 0}

    def download_file(file_id):
        calls["download"] += 1
        return xlsx_bytes

    excel_photos.sync_cache(cache_dir=tmp_path, list_files=lambda: files, download_file=download_file)
    summary2 = excel_photos.sync_cache(cache_dir=tmp_path, list_files=lambda: files, download_file=download_file)

    assert calls["download"] == 1
    assert summary2 == {"processed": 0, "skipped": 1, "photos": 0}

    index_path = tmp_path / "index.json"
    entries = json.loads(index_path.read_text(encoding="utf-8"))
    assert len(entries) == 1


def test_sync_cache_reprocesses_and_replaces_entries_when_file_updated(tmp_path):
    png1 = _png_bytes((255, 0, 0))
    wb1 = _build_workbook({"sheet1": [png1]})
    out1 = BytesIO()
    wb1.save(out1)

    png2_a = _png_bytes((0, 255, 0))
    png2_b = _png_bytes((0, 0, 255))
    wb2 = _build_workbook({"sheet1": [png2_a, png2_b]})
    out2 = BytesIO()
    wb2.save(out2)

    name = "(L1)2026.03.10_0852_A品種_T5.xlsx"
    state = {"updated_at": "2026-03-10T08:52:00", "content": out1.getvalue()}

    list_files = lambda: [{"id": "file1", "name": name, "updated_at": state["updated_at"]}]
    download_file = lambda file_id: state["content"]

    excel_photos.sync_cache(cache_dir=tmp_path, list_files=list_files, download_file=download_file)

    state["updated_at"] = "2026-03-11T09:00:00"
    state["content"] = out2.getvalue()
    summary2 = excel_photos.sync_cache(cache_dir=tmp_path, list_files=list_files, download_file=download_file)

    assert summary2 == {"processed": 1, "skipped": 0, "photos": 2}

    index_path = tmp_path / "index.json"
    entries = json.loads(index_path.read_text(encoding="utf-8"))
    assert len(entries) == 2
