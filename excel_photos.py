# -*- coding: utf-8 -*-
"""
Excel埋め込み写真の抽出モジュール
==================================

【背景】
現場写真はBox上のExcelファイル内に埋め込まれていることがほとんど。
box_client.py はフォルダ内の画像ファイル直置きを前提としているため、
Excelから画像を抽出する本モジュールを別途用意している。

【確定している要件】
- 写真と日時の対応は「ファイル名・シート名に日付が入っているのみ」
  → 写真の時刻粒度は日単位
- ファイル名規則: config.EXCEL_FILENAME_PATTERN(命名規則: (ライン名)YYYY.MM.DD_HHMM_品種_厚みコード.xlsx)
- Excelファイルの分割単位は不規則(1日1ファイルとは限らない)
- 処理方式は「定期的な事前抽出+キャッシュ」+「今すぐ同期」ボタンの
  ハイブリッド(ユーザーと合意済み)

【実装方針】
1. 同期処理(sync_cache):
   - Boxの対象フォルダから .xlsx/.xls ファイル一覧を取得
   - ファイル名が命名規則に一致しないものはスキップ
   - manifest.json(file_id→updated_at)と比較し、Box側で更新されていない
     ファイルはスキップ(差分同期)
   - openpyxl でロードし、_extract_images_from_workbook で埋め込み画像を抽出
   - 画像は cache/photos/{date}/{hash}.{拡張子} に保存し、
     cache/index.json に get_photos が読める形式で記録
   - 更新されたファイルは、同じsource_file_idを持つ古いindexエントリを
     差し替える
   - ※ .xls (旧形式) が混在する場合はopenpyxlでは開けない。
     既存のxls→xlsx変換スクリプト(xlrd+openpyxl)の流用を検討
2. 参照処理(get_photos):
   - cache/index.json を読み、日付範囲でフィルタして返す
   - box_client.get_photos と同じ戻り値形式に合わせ、app.py側の
     切り替えを最小限にする
3. UI(未実装):
   - フィルターパネルに「写真を今すぐ同期」ボタンを追加
   - 同期は時間がかかるため、POST /api/photos/sync を非同期実行し
     進捗をポーリングするか、完了までスピナー表示
"""


import hashlib
import json
import re
from datetime import date as _date
from datetime import datetime as _datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

import box_client
import config

_FILENAME_RE = re.compile(config.EXCEL_FILENAME_PATTERN)
_INDEX_PATH = Path(config.EXCEL_PHOTO_CACHE_DIR) / "index.json"


def parse_filename(filename: str):
    """
    Excelファイル名からライン名と日付を抽出する。

    命名規則(config.EXCEL_FILENAME_PATTERN)に一致しない場合、または
    日付として不正な場合は None を返す(例外は投げない。sync_cache側で
    対象外ファイルとしてスキップする用途を想定)。

    Returns
    -------
    dict {"line_name": str, "date": date, "filename": str} | None
    """
    m = _FILENAME_RE.search(filename)
    if not m:
        return None

    groups = m.groupdict()
    try:
        parsed_date = _date(int(groups["year"]), int(groups["month"]), int(groups["day"]))
    except ValueError:
        return None

    return {
        "line_name": groups.get("line_name", ""),
        "date": parsed_date,
        "filename": filename,
    }


def _extract_images_from_workbook(wb):
    """
    ワークブック内の全シートから埋め込み画像を抽出する。

    Returns
    -------
    list[tuple[str, bytes, str]]  # [(シート名, 画像バイト列, 拡張子), ...]
    """
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for image in getattr(ws, "_images", []):
            results.append((sheet_name, image._data(), image.format))
    return results


def _load_json(path, default):
    if not path.exists():
        return default
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _save_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def sync_cache(cache_dir=None, list_files=None, download_file=None):
    """
    Box上のExcelファイルから写真を抽出し、ローカルキャッシュを更新する。

    命名規則(config.EXCEL_FILENAME_PATTERN)に一致しないファイル、および
    前回同期時から更新されていないファイル(manifest.jsonで判定)はスキップする。

    Parameters
    ----------
    cache_dir : キャッシュ保存先。省略時はconfig.EXCEL_PHOTO_CACHE_DIR
    list_files : () -> list[dict]。省略時はBox上のconfig.EXCEL_BOX_FOLDER_IDを参照
    download_file : (file_id) -> bytes。省略時はbox_client.download_file_content

    Returns
    -------
    dict {"processed": int, "skipped": int, "photos": int}
    """
    base_dir = Path(cache_dir) if cache_dir is not None else Path(config.EXCEL_PHOTO_CACHE_DIR)
    index_path = base_dir / "index.json"
    manifest_path = base_dir / "manifest.json"

    if list_files is None:
        list_files = lambda: box_client.list_excel_files(config.EXCEL_BOX_FOLDER_ID)
    if download_file is None:
        download_file = box_client.download_file_content

    manifest = _load_json(manifest_path, {})
    index_entries = _load_json(index_path, [])

    processed = 0
    skipped = 0
    photos_added = 0

    for file_info in list_files():
        parsed = parse_filename(file_info["name"])
        if parsed is None:
            skipped += 1
            continue

        file_id = file_info["id"]
        updated_at = file_info.get("updated_at")
        if manifest.get(file_id, {}).get("updated_at") == updated_at:
            skipped += 1
            continue

        content = download_file(file_id)
        wb = load_workbook(BytesIO(content))
        images = _extract_images_from_workbook(wb)

        # 同じソースファイルの古いエントリは差し替える
        index_entries = [e for e in index_entries if e.get("source_file_id") != file_id]

        date_str = parsed["date"].isoformat()
        photo_dir = base_dir / "photos" / date_str

        for sheet_name, image_bytes, fmt in images:
            image_hash = hashlib.sha256(image_bytes).hexdigest()[:16]
            photo_dir.mkdir(parents=True, exist_ok=True)
            image_path = photo_dir / f"{image_hash}.{fmt}"
            with open(image_path, "wb") as img_f:
                img_f.write(image_bytes)

            index_entries.append({
                "id": image_hash,
                "name": f"{file_info['name']}#{sheet_name}",
                "date": date_str,
                "line_name": parsed["line_name"],
                "path": str(image_path),
                "source_file_id": file_id,
            })
            photos_added += 1

        manifest[file_id] = {"updated_at": updated_at, "name": file_info["name"]}
        processed += 1

    _save_json(index_path, index_entries)
    _save_json(manifest_path, manifest)

    return {"processed": processed, "skipped": skipped, "photos": photos_added}


def get_photos(start, end, index_path=None):
    """
    キャッシュ(index.json)から、日付範囲 [start, end] に入る写真を返す。

    box_client.get_photos と同じ戻り値形式に合わせる。写真の時刻粒度は
    日単位のため、timestampは各写真の日付の00:00:00とする。

    Returns
    -------
    list[dict]  # [{id, name, timestamp}, ...] timestamp昇順
    """
    path = Path(index_path) if index_path is not None else _INDEX_PATH
    if not path.exists():
        return []

    with open(path, encoding="utf-8") as f:
        entries = json.load(f)

    start_date = start.date() if isinstance(start, _datetime) else start
    end_date = end.date() if isinstance(end, _datetime) else end

    results = []
    for entry in entries:
        entry_date = _date.fromisoformat(entry["date"])
        if start_date <= entry_date <= end_date:
            results.append({
                "id": entry["id"],
                "name": entry["name"],
                "timestamp": _datetime.combine(entry_date, _datetime.min.time()).isoformat(),
            })

    results.sort(key=lambda x: x["timestamp"])
    return results
